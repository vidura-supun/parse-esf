import json, os, sys
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FONT  = Font(bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def extract_paths(obj, found=None):
    if found is None: found = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            if k == "path" and isinstance(v, str) and v:
                found.append(v)
            else:
                extract_paths(v, found)
    elif isinstance(obj, list):
        for item in obj: extract_paths(item, found)
    return found


def parse_time(raw_time):
    """Parse ESF ISO 8601 timestamp (may have nanoseconds) to (utc_str, local_str)."""
    try:
        if not raw_time:
            return "", ""
        ts = raw_time.rstrip("Z")
        if "." in ts:
            base, frac = ts.split(".", 1)
            ts = base + "." + frac[:6]
        dt_utc = datetime.fromisoformat(ts).replace(tzinfo=timezone.utc)
        return (
            dt_utc.strftime("%Y-%m-%d %H:%M:%S.%f UTC"),
            dt_utc.astimezone().strftime("%Y-%m-%d %H:%M:%S.%f %Z"),
        )
    except (ValueError, TypeError):
        return "", ""


def style_header_row(ws, fields, display_names=None):
    labels = display_names if display_names else fields
    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font  = HEADER_FONT
        cell.fill  = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def autosize_columns(ws, max_width=80):
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, max_width)


def build_process_tree(raw_events):
    """
    Build a process tree from fork and exec events.

    - fork  → records a new child process (pid, ppid, exe, start_time)
    - exec  → updates an existing pid with new exe and cmdline
    - All 'process' fields are seeded as known processes (captures roots).

    Returns a list of dicts ordered by DFS walk, each with:
        pid, ppid, depth, exe, cmdline, start_time
    """
    processes = {}  # pid -> dict

    def upsert(pid, ppid, exe, cmdline, start_time):
        if pid is None:
            return
        if pid not in processes:
            processes[pid] = {"pid": pid, "ppid": ppid, "exe": exe,
                              "cmdline": cmdline, "start_time": start_time}
        else:
            existing = processes[pid]
            if ppid    != "": existing["ppid"]       = ppid
            if exe     != "": existing["exe"]        = exe
            if cmdline != "": existing["cmdline"]    = cmdline
            if start_time != "": existing["start_time"] = start_time

    # Pass 1: seed known processes from the 'process' field of every event (establishes roots)
    for event in raw_events:
        proc  = event.get("process", {})
        audit = proc.get("audit_token", {})
        upsert(
            pid        = audit.get("pid"),
            ppid       = proc.get("ppid", ""),
            exe        = proc.get("executable", {}).get("path", ""),
            cmdline    = "",
            start_time = proc.get("start_time", ""),
        )

    # Pass 2: fork/exec events override with authoritative child/target data
    for event in raw_events:
        evt_data = event.get("event", {})
        key = list(evt_data.keys())[0] if evt_data else None

        if key == "fork":
            child   = evt_data["fork"].get("child", {})
            c_audit = child.get("audit_token", {})
            upsert(
                pid        = c_audit.get("pid"),
                ppid       = child.get("ppid", ""),
                exe        = child.get("executable", {}).get("path", ""),
                cmdline    = "",
                start_time = child.get("start_time", ""),
            )

        elif key == "exec":
            target  = evt_data["exec"].get("target", {})
            t_audit = target.get("audit_token", {})
            args    = evt_data["exec"].get("args", [])
            upsert(
                pid        = t_audit.get("pid"),
                ppid       = target.get("ppid", ""),
                exe        = target.get("executable", {}).get("path", ""),
                cmdline    = " ".join(args) if args else "",
                start_time = target.get("start_time", ""),
            )

    # Build children map
    children = {}
    all_pids = set(processes)
    for pid, proc in processes.items():
        ppid = proc["ppid"]
        children.setdefault(ppid, []).append(pid)

    # Roots: processes whose ppid is not a known pid
    roots = sorted(pid for pid, proc in processes.items() if proc["ppid"] not in all_pids)

    # Iterative DFS to avoid recursion limits on deep trees
    tree_rows = []
    stack = [(pid, 0) for pid in reversed(roots)]
    while stack:
        pid, depth = stack.pop()
        proc = processes[pid]
        tree_rows.append({
            "pid":        pid,
            "ppid":       proc["ppid"],
            "depth":      depth,
            "exe":        proc["exe"],
            "cmdline":    proc["cmdline"],
            "start_time": proc["start_time"],
        })
        for child_pid in reversed(sorted(children.get(pid, []))):
            stack.append((child_pid, depth + 1))

    return tree_rows


def parse_esf_jsonl(input_path, output_path):
    rows = []
    raw_events = []

    with open(input_path, "rb") as f:
        raw = f.read()
    encoding = "utf-16" if raw[:2] in (b'\xff\xfe', b'\xfe\xff') else "utf-8"
    content  = raw.decode(encoding).lstrip('\ufeff')

    decoder = json.JSONDecoder()
    pos = 0
    while pos < len(content):
        while pos < len(content) and content[pos] in ' \t\r\n':
            pos += 1
        if pos >= len(content):
            break
        try:
            event, pos = decoder.raw_decode(content, pos)
        except json.JSONDecodeError:
            if content[pos:].lstrip().startswith('{'):
                break  # truncated final event
            pos += 1
            continue

        raw_events.append(event)

        proc         = event.get("process", {})
        audit        = proc.get("audit_token", {})
        evt_data     = event.get("event", {})
        action       = event.get("action", {})
        result       = action.get("result", {})
        result_inner = result.get("result", {})
        paths        = extract_paths(evt_data)

        exec_data = evt_data.get("exec", {})
        args      = exec_data.get("args", [])
        cmdline   = " ".join(args) if args else ""

        date_time_utc, date_time_local = parse_time(event.get("time", ""))

        rows.append({
            "date_time_local":            date_time_local,
            "date_time_utc":              date_time_utc,
            "mach_time":                  event.get("mach_time", ""),
            "global_seq_num":             event.get("global_seq_num", ""),
            "seq_num":                    event.get("seq_num", ""),
            "event_type_id":              event.get("event_type"),
            "action_type":                event.get("action_type", ""),
            "result_type":                result.get("result_type", ""),
            "result_auth":                result_inner.get("auth", ""),
            "result_flags":               result_inner.get("flags", ""),
            "process_pid":                audit.get("pid", ""),
            "process_ppid":               proc.get("ppid", ""),
            "process_uid":                audit.get("ruid", ""),
            "process_gid":                audit.get("rgid", ""),
            "process_euid":               audit.get("euid", ""),
            "process_session_id":         proc.get("session_id", ""),
            "process_exe":                proc.get("executable", {}).get("path", ""),
            "process_signing_id":         proc.get("signing_id", ""),
            "process_team_id":            proc.get("team_id", ""),
            "process_is_platform_binary": proc.get("is_platform_binary", ""),
            "process_codesigning_flags":  proc.get("codesigning_flags", ""),
            "process_start_time":         proc.get("start_time", ""),
            "event_key":                  list(evt_data.keys())[0] if evt_data else "",
            "cmdline":                    cmdline,
            "paths":                      " | ".join(paths) if paths else "",
            "raw_event":                  json.dumps(evt_data),
        })

    if not rows:
        print("No events parsed.")
        return

    priority   = ["date_time_local", "date_time_utc", "process_pid", "process_start_time",
                  "process_exe", "event_key", "cmdline", "paths", "raw_event"]
    rest       = [k for k in rows[0] if k not in priority]
    fieldnames = priority + rest

    tree_rows = build_process_tree(raw_events)

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # Sheet 1: Events
    ws_events       = wb.active
    ws_events.title = "Events"
    style_header_row(ws_events, fieldnames)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, field in enumerate(fieldnames, 1):
            ws_events.cell(row=row_idx, column=col_idx, value=str(row.get(field, "")))

    ws_events.freeze_panes = "A2"
    ws_events.auto_filter.ref = ws_events.dimensions
    autosize_columns(ws_events)

    # Sheet 2: Process Tree
    ws_tree       = wb.create_sheet("Process Tree")
    tree_fields   = ["pid", "ppid", "depth", "exe", "cmdline", "start_time"]
    tree_labels   = ["PID", "PPID", "Depth", "Executable", "Command Line", "Start Time"]
    style_header_row(ws_tree, tree_fields, display_names=tree_labels)

    for row_idx, proc in enumerate(tree_rows, 2):
        depth = proc["depth"]
        for col_idx, field in enumerate(tree_fields, 1):
            val = str(proc.get(field, ""))
            if field == "exe" and depth > 0:
                val = "  " * depth + "\u2514\u2500 " + val
            ws_tree.cell(row=row_idx, column=col_idx, value=val)

    ws_tree.freeze_panes = "A2"
    autosize_columns(ws_tree)

    wb.save(output_path)
    print(f"Done. {len(rows)} events | {len(tree_rows)} processes | {output_path}")


if __name__ == "__main__":
    if len(sys.argv) == 2:
        input_path  = sys.argv[1]
        output_path = os.path.splitext(input_path)[0] + ".xlsx"
        parse_esf_jsonl(input_path, output_path)
    elif len(sys.argv) == 3:
        parse_esf_jsonl(sys.argv[1], sys.argv[2])
    else:
        print("Usage: parse_esf.py <input> [output.xlsx]")
        sys.exit(1)
